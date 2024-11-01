import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
import re

# 결전 변수 및 스타일 정의
first_file_path, first_file_data = None, None
yellow_fill, blue_fill = PatternFill(start_color="FFFF00", fill_type="solid"), PatternFill(start_color="87CEEB", fill_type="solid")
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

shop_mapping = {
        "ESM옥션": "A",
        "GS shop": "GS",
        "11번가": "11",
        "CJ온스타일": "CJ",
        "ESM지마켓": "G",
        "롯데온": "ON",
        #"올웨이즈": "올웨이즈",
        "센트럴리빙": "자-벤",
        "쿠팡": "쿠팡-벤",
        "신세계몰(신)": "SSG",
        "현대이지웰": "이지웰",
        "패션플러스": "패플",
        "도매꾹":"도",
        #"텐바이텐":"텐바이텐",
        "LF몰":"Lfmall",
        "카카오톡스토어":"카카오톡",
        #"29CM":"29CM",
        "브랜디":"하이버",
        "shop by":"티딜"
}

keywords_to_remove = [
    r"사이즈 선택:",r"향 선택:", r"모델명:",r"수량:","차량용", "벤오토", "센트럴리빙", "컵홀드", "종류","와이퍼", r"제품:\d+\.?", r"\(\+\d+원\)", r"\(\-\d+원\)", r"\d+-\d+\.",r"\(\d+원\)",
    r"\[한정수량\]", r"선택:", r"선택 :",r"선택\d+:",r"선택\.\d+", r"색상:",r"컬러:", r"(?<!ml)\b\d+\.\b(?!ml)" ,r"\/\-\d+원\/",r"\/\+\d+원\/",r"\/\d+원\/",
    r"\/",r"\:", r"\,", r"옵션=",r"선택=",r"\[국내배송\]",r"선택\s+\d", r"\(\+\d+\)","mm",r"색상=",r"사이즈=","목베개 목받침 헤드레스트 메모리폼","자동차",
    r"브랜드=", r"모델명=","차종",r"추가:",r"추가 :","코란도"
]

keyword_replacements = {
    "^": " ", "자동차 초미세모  세차":" ", "자동차 초미세모 세차":" ", "운전석 조수석 세트": " ",r"[1+1 원쁠]":" ","자동차 청소 디테일링 초미세모":" ",
    r"(1대분은 ★2개★ 구매바랍니다.)":"", r"=(1103~1506)":" ",

    r"（K5）":" ", r"（쏘나타）":" ", r"（그랜저）":" ",

    r"카니발 4세대 KA4 （20~）":" ",r"그랜드카니발 R (05~14)":" ",r"그랜드카니발 R （05~14）":" ",r"카니발 12 후방 （98~05）":" ",
    r"카니발 12 （98~05）":" ",r"뉴카니발 R （06~14）":" ",r"카니발 올뉴더뉴 （~20）":" ",r"카니발 KA4 PE （23~）":" ",
    r"카니발 4세대 KA4 (20~)":" ",r"그랜드카니발 R (05~14)":" ",r"카니발 12 후방 (98~05)":" ",
    r"카니발 12 (98~05)":" ",r"뉴카니발 R (06~14)":" ",r"카니발 올뉴더뉴 (~20)":" ",r"카니발 KA4 PE (23~)":" ",

    r"G330 G380（14~16년）":" ", r"G330 G380(14~16년)":" ",

    r"쏘렌토 MQ4（20년~）":" ", r"쏘렌토 MQ4(20년~)":" ",

    r"LF 뉴라이즈（14 3~19）":" ", r"LF뉴라이즈(143~19)": " ", r"YF쏘나타(0909)":" ", r"NF쏘나타(049~098)":" ",r"LF 뉴라이즈（143~19）":" ",

    r"스타리아(21년~)": " ", r"스타리아 （21년~）":" ",

    r"그랜저 HG （11 01~）":" ", r"더뉴그랜저IG （20~）":" ", r"그랜저 HG (1101~)":" ",r"더뉴그랜저IG (20~)":" ",

    r"올뉴모닝(1701~)": " ", r"올뉴모닝(1701~)": " ",r"모닝올뉴모닝(04~16)":" ",

    r"투싼NX4 (2009~)":" ",r"올뉴투싼 (1503~)":" ",r"투싼ix （0908~）":" ",r"투싼 （0404~0907）":" ",r"투싼ix (0908~)":" ",r"투싼 (0404~0907)":" ",

    r"SM3 뉴 (098~14년)":" ",r"SM3 (05~09년)":" ",r"SM5 뉴플래티넘노바": " ",
    
    r"더뉴 QM6 (19년 ~ 현재)":" ",

    r"팰리세이드 (19년~)": " ", r"팰리세이드 （19년~）":" ",

    r"스포티지 NQ5（21년）":" ", r"스포티지 후방 단품 （93~02년）":" ", r"스포티지 （93~02년）":" ", r"스포티지 뉴（048~103）":" ",
    r"스포티지 R（104~158）":" ", r"올뉴더볼드（1509）":" ", r"스포티지 R（10 4~15 8）":" ",

    #코란도
    r"C (11~18년)": " ",r"뷰티풀（19년~）":" ", r"투리스모（12~19년）":" ", r"스포츠（12~18년）":" ", r"C （11~18년）":" ", r"뉴（96~04년）":" ",
    
    r"DM더스타일(05년~12년)": " ", r"DM더프라임（12년~18년）":" ",
    
    r"토레스":" ",

    r"아반떼 HD(0605)": " ", r"아반떼 HD（0605）":" ", r"아반떼XD （004~065）":" ", r"아반떼 XD 후방 단품":" ", r"아반떼 MD쿠페 （1008）":" ",
    r"아반떼 AD （1509）": " ", r"더뉴아반떼（1809）":" ", r"아반떼 cn7（2004~）":" ",
    r"아반떼 HD(0605)":" ", r"아반떼XD (004~065)":" ",r"아반떼 MD쿠페 (1008)":" ",
    r"아반떼 AD (1509)": " ", r"더뉴아반떼(1809)":" ", r"아반떼 cn7(2004~)":" ",
    
    r"포르테쿱": " ",

    r"더뉴 레이":" ",

    r"셀토스더뉴셀토스":" ",

    r"봉고3":" ", r"봉고2":" ",

    r"올뉴마이티":" ", r"마이티 II":" ",r"마이티 I （8604~）":" ",r"마이티 I (8604~)":" ",
    
    r"더뉴그랜드(18년~)":" ", r"그랜드 （0705~18년）":" ",r"그랜드 (0705~18년)":" ",

    r"베라크루즈 (11년~15년)":" ",r"베라크루즈 뉴（11년~15년）":" ",r"베라크루즈 뉴(11년~15년)":" ",

    r"카운티":" ", r"카운티 후방 단품":" ",

    r"올뉴말리부(1604~)":" ", r"말리부（~1603）":" ", r"말리부(~1603)":" ",

    r"더뉴K3 （204）":" ", r"올뉴K3 （183~204）":" ", r"더뉴K3 유로쿱 （15~18）":" ", r"K3 유로쿱 （12~16）":" ",
    r"더뉴K3 (204)":" ", r"올뉴K3 (183~204)":" ", r"더뉴K3 유로쿱 (15~18)":" ", r"K3 유로쿱 (12~16)":" ",

    r"K5 DL3 （20~）":" ", r"더뉴 K5 （104~156）":" ", r"올뉴 K5 （157~19）":" ",
    r"K5 DL3 (20~)":" ", r"더뉴 K5 (104~156)":" ", r"올뉴 K5 (157~19)":" ",
    
    r"K7 더뉴올뉴 （12~19）":" ", r"K7 더프리스티지 （09~12）":" ", r"K7 프리미어 （196）":" ",
    r"K7 더뉴올뉴 (12~19)":" ", r"K7 더프리스티지 (09~12)":" ", r"K7 프리미어 (196)":" ",

    r"K8 GL3 PE (24~)":" ",
    
    r"K9더뉴K9 （12~18）":" ", r"더K9 （1805）":" ",r"K9더뉴K9 (12~18)":" ", r"더K9 (1805)":" "
}

keywords_to_find = [
    "듀오세트", "앤코코", "수세미","센트세이", "석고"
]

# 중복 처리를 할 키워드 리스트 추가
duplicate_keywords = [
    "수세미", "앤코코", "듀오세트","센트세이","먼지털이 미니 브러쉬","컵홀더","타이탄 청소기 헤파필터", "석고"
]  # 원하는 키워드를 추가

def convert_shop_name(shop_name, product_name, option):
    # 상품명이 NaN이거나 문자열이 아닌 경우 빈 문자열로 변환
    if not isinstance(product_name, str):
        product_name = "" if pd.isna(product_name) or isinstance(product_name, (float, int)) else str(product_name)
    if not isinstance(option, str):
        option = "" if pd.isna(option) or isinstance(option, (float, int)) else str(option)

    if shop_name == "스마트스토어":
        # 상품명과 옵션에서 키워드가 있는지 확인
        if any(keyword in product_name or keyword in option for keyword in [
            "벤오토", "운전석", "조수석", "스트랩", "각인", "브러쉬", "카우하이드",r"추가:",
            "키체인", "키링", "2in1", "가죽", "와이퍼", "듄", "메모리폼", "석고방향제",
            "트렁크", "틈새가죽", "차종", "화이트골드", "틈새", "아이온", "일랑일랑",
            "딥 블랙", "엣지형", "블랙골드", "햇빛가리개", "정리광","콘솔","3열추가",
            "도어가드", "사이즈", "포맥스", "컵홀더", "콘솔수납포켓","타이탄 청소기"
        ]): 
            return "N-벤"
        else:
            return "N-올뷰"

    return shop_mapping.get(shop_name, shop_name)

def clean_option(option):
    # "옵션명 [해당내용]" 형식 제거
    option = re.sub(r"옵션명\s*\[.*?\]", "", option)

    # 지정된 패턴을 모두 ★\d개★로 변경 (단, 반사스티커가 포함된 경우는 제외)
    if "반사스티커" not in option:
        patterns = [
            r"x\s*(\d+)\s*개(?!입)",      # "x 2개" 뒤에 "입"이 없을 때만 별로 감싸기
            r"\((\d+)\s*개(?!입)\)",      # "(2개)" 뒤에 "입"이 없을 때만 별로 감싸기
            r"-\s*(\d+)\s*개(?!입)",      # "-2개" 뒤에 "입"이 없을 때만 별로 감싸기
            r"(\d+)\s*개(?!입)",          # "2개" 뒤에 "입"이 없을 때만 별로 감싸기
            r"\((\d+)\s*롤\)",            # "(2롤)" -> "★2롤★"
            r"(\d+)\s*롤",                # "2롤" -> "★2롤★"
            r"\(총\s*(\d+)\s*롤\)",       # "(총 2롤)" -> "★2롤★"
            r"\(총\s*(\d+)\s*개(?!입)\)", # "(총 2개)" 뒤에 "입"이 없을 때만 별로 감싸기
            r"총\s*(\d+)\s*개(?!입)",     # "총 2개" 뒤에 "입"이 없을 때만 별로 감싸기
        ]

        for pattern in patterns:
            option = re.sub(pattern, r"★\1개★", option)

        # 괄호 안에 ★\d개★가 있으면 괄호를 제거하고 그 안의 ★\d개★만 유지
        option = re.sub(r"[()（）【】](총\s*)?(\u2605\d+개\u2605)[()（）【】]", r"\2", option)

    # 반사스티커가 포함되어 있는 경우 개수를 그대로 유지
    for keyword in keywords_to_remove:
        option = re.sub(keyword, "", option)
    
    # 슬래시 및 전각 슬래시 직접 제거
    option = option.replace("/", " ").replace("／", " ")
    
    for key, value in keyword_replacements.items():
        if key in option:
            option = option.replace(key, value)
    
    return option.strip()

def adjust_quantity_for_mini_brush(option):
    if "미니 브러쉬" in option:
        match = re.search(r"★(\d+)개★", option)
        if match:
            quantity = int(match.group(1))
            if quantity > 1:
                adjusted_quantity = quantity // 2
                if adjusted_quantity >= 1:
                    option = re.sub(r"★\d+개★", f" ★{adjusted_quantity}개★", option)
                else:
                    option = re.sub(r"★\d+개★", "", option)
    return option

def find_keyword(product_name):
    for keyword in keywords_to_find:
        if keyword in product_name:
            if keyword == "앤코코":
                return f"[{keyword}]"
            else:
                return keyword
    return ""

def open_first_file():
    global first_file_path, first_file_data
    first_file_path = filedialog.askopenfilename(title="첫 번째 엑셀 파일 선택", filetypes=[["엑셀 파일", "*.xlsx"]])
    if first_file_path:
        first_file_data = pd.read_excel(first_file_path)
        messagebox.showinfo("성공", "첫 번째 파일이 성공적으로 읽히였습니다.")

def open_second_file():
    if first_file_data is None or first_file_data.empty:
        messagebox.showwarning("경고", "먼저 첫 번째 파일을 선택하세요.")
        return
    second_file_path = filedialog.askopenfilename(title="두 번째 엑셀 파일 선택", filetypes=[["엑셀 파일", "*.xlsx"]])
    if second_file_path:
        process_files(second_file_path)

def process_files(second_file_path):
    try:
        df2 = pd.read_excel(second_file_path)

         # 첫 번째 파일의 '옵션'에 '[듀오세트]'가 있는 경우 삭제
        first_file_data['옵션'] = first_file_data['옵션'].str.replace(r"\[듀오세트\]", "", regex=True)
        first_file_data['옵션'] = first_file_data['옵션'].str.replace(r"석고방향제", "", regex=True)
        
        # 모든 행의 데이터를 안전하게 문자열로 변환
        df2['품목명'] = first_file_data.apply(
            lambda row: f"({convert_shop_name(row['쇼핑몰'], row['상품명'], str(row['옵션']) if not pd.isna(row['옵션']) else '')}) {find_keyword(row['상품명'])} {adjust_quantity_for_mini_brush(clean_option(str(row['옵션']) if not pd.isna(row['옵션']) else ''))}" + (f"  ♥{row['주문수량']}개♥" if row['주문수량'] > 1 else ""), 
            axis=1
        )
        
        # 중복 키워드 처리: 'duplicate_keywords' 리스트에 있는 키워드만 중복을 제거
        for keyword in duplicate_keywords:
            df2['품목명'] = df2['품목명'].apply(lambda x: re.sub(rf"({keyword})", r"\1", x, count=1))  # 첫 번째만 남기기
            df2['품목명'] = df2['품목명'].apply(lambda x: re.sub(rf"({keyword}).*?\1", r"\1", x))  # 중복된 부분을 제거
        # 품목명에서 6자리 연속된 숫자가 있는 경우 3자리에서 띄어쓰기 추가
        df2['품목명'] = df2['품목명'].str.replace(r'(\d{3})(\d{3})', r'\1 \2', regex=True)
        df2['받는분성명'] = first_file_data['받는분성함']
        df2[['받는분전화번호', '받는분기타연락처', '받는분주소(전체 분할)', '배송메세지1']] = first_file_data[['반는분연락처', '받는분기타연락처', '받는분주소', '배송메시지']]

        # 중복 주소 처리: 같은 주소에 대해 첫 번째 품목명에만 쇼핑몰 이름을 나타내고 나머지는 빈 값으로 설정
        previous_address = None
        previous_shop_name = None  # 이전 행의 쇼핑몰 이름을 저장

        for i in range(len(df2)):
            current_address = df2.at[i, '받는분주소(전체 분할)']  # '받는분주소(전체 분할)' 열 기준으로 주소 확인
            current_shop_name = convert_shop_name(first_file_data.at[i, '쇼핑몰'], first_file_data.at[i, '상품명'], first_file_data.at[i, '옵션'])  # 현재 행의 쇼핑몰 이름 확인

            if current_address == previous_address and current_shop_name == previous_shop_name:
                # 중복된 행에서는 동일한 쇼핑몰 이름 제거
                df2.at[i, '품목명'] = re.sub(rf"\({re.escape(current_shop_name)}\)", "", df2.at[i, '품목명']).strip()
            
            previous_address = current_address  # 이전 주소 업데이트
            previous_shop_name = current_shop_name  # 이전 쇼핑몰 이름 업데이트

        # 품목명에서 ★1개★가 존재하면 삭제
        df2['품목명'] = df2['품목명'].str.replace(r'★1개★', '', regex=True)

        # 연속된 ★을 하나로 합치기 (★★ => ★)
        df2['품목명'] = df2['품목명'].str.replace(r'★{2,}', '★', regex=True)

        # 품목명에 별이 1개만 있을 경우 삭제
        df2['품목명'] = df2['품목명'].apply(lambda x: x.replace('★', '') if x.count('★') == 1 else x)
        
        # '22.'와 같은 숫자 제거
        df2['품목명'] = df2['품목명'].apply(lambda x: re.sub(r"\b\d+\.\s*", "", x))  # 숫자. 제거

        # 엑셀 파일 저장
        new_file_path = generate_new_file_path(second_file_path)
        df2.to_excel(new_file_path, index=False)
        apply_styles(new_file_path)
        messagebox.showinfo("성공", f"파일이 저장되었습니다: {new_file_path}")
    except Exception as e:
        messagebox.showerror("오류", f"파일을 처리하는 중 문제가 발생했습니다: {e}")

def generate_new_file_path(second_file_path):
    today = datetime.now().strftime('%Y%m%d')
    save_dir = os.path.dirname(second_file_path)
    new_file_path = os.path.join(save_dir, f"{today}.xlsx")
    counter = 1
    while os.path.exists(new_file_path):
        new_file_path = os.path.join(save_dir, f"{today}_{counter}.xlsx")
        counter += 1
    return new_file_path

def apply_styles(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    yellow_columns, blue_columns, address_column = ['C', 'D', 'E', 'G', 'I', 'J', 'K', 'N'], ['C', 'D', 'E', 'G', 'I', 'J', 'K', 'N'], 'G'
    previous_address = None
    
    for row in range(1, ws.max_row + 1):
        for col in yellow_columns:
            cell = ws[f"{col}{row}"]
            if row == 1:
                cell.fill = yellow_fill
            if cell.value:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True)
        current_address = ws[f"{address_column}{row}"].value
        if current_address and current_address == previous_address:
            for col in blue_columns:
                ws[f"{col}{row}"].fill = blue_fill
                ws[f"{col}{row - 1}"].fill = blue_fill
        previous_address = current_address

    for col in ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2

    ws.auto_filter.ref = ws.dimensions
    wb.save(file_path)

root = tk.Tk()
root.title("J CORE PROJECT")
root.geometry("400x250")
root.configure(bg="#2C3E50")
root.resizable(False, False)  # 창 크기 고정

label = tk.Label(root, text="엑셀 파일을 선택하세요:", font=("Helvetica", 16), fg="#ECF0F1", bg="#2C3E50")
label.pack(pady=20)

first_button = tk.Button(root, text="첫 번째 파일 선택", command=open_first_file, font=("Helvetica", 14), bg="#3498DB", fg="white", activebackground="#2980B9", activeforeground="white")
first_button.pack(pady=10, ipadx=10, ipady=5)

second_button = tk.Button(root, text="두 번째 파일 선택", command=open_second_file, font=("Helvetica", 14), bg="#E74C3C", fg="white", activebackground="#C0392B", activeforeground="white")
second_button.pack(pady=10, ipadx=10, ipady=5)

root.mainloop()