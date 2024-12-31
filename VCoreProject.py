import tkinter as tk
from tkinter import font, scrolledtext
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import threading
import openpyxl
import datetime
import re
import logging
import os
import shutil
from webdriver_manager.chrome import ChromeDriverManager
import ctypes
import sys

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# 파일 상단에 상수 정의
CHROME_DRIVER_PATH = r"C:\Program Files\SeleniumBasic\chromedriver.exe"
USER_DATA_DIR = r"C:\VCP"
LOGIN_EMAIL = "marketing11111111111@gmail.com"
TARGET_DIR = r"C:\Program Files\SeleniumBasic"

# 관리자 권한 확인 함수
def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

# 확인 메시지 창 표시 함수
def show_message(title, text):
    ctypes.windll.user32.MessageBoxW(0, text, title, 0x40)

# 크롬 드라이버 자동 업데이트 함수
def update_chrome_driver():
    try:
        # webdriver-manager로 크롬 드라이버 다운로드
        driver_path = ChromeDriverManager().install()
        print(f"다운로드된 드라이버 위치: {driver_path}")

        # 지정된 경로로 복사 준비
        if not os.path.exists(TARGET_DIR):
            os.makedirs(TARGET_DIR)

        target_path = os.path.join(TARGET_DIR, "chromedriver.exe")

        # 기존 파일이 있으면 삭제
        if os.path.exists(target_path):
            os.remove(target_path)
            print(f"기존 드라이버 파일을 삭제했습니다: {target_path}")

        # 새 드라이버 복사
        shutil.copy(driver_path, target_path)
        print(f"크롬 드라이버가 성공적으로 업데이트되었습니다! 위치: {target_path}")

        # 확인 메시지 표시
        show_message("업데이트 완료", f"크롬 드라이버가 성공적으로 업데이트되었습니다.\n위치: {target_path}")

    except Exception as e:
        error_message = f"드라이버 업데이트 중 오류가 발생했습니다:\n{e}"
        print(error_message)
        show_message("업데이트 오류", error_message)

# 드라이버 업데이트 실행 함수
def on_update_driver():
    if not is_admin():
        # 관리자 권한으로 스크립트를 재실행
        print("관리자 권한이 필요합니다. 프로그램을 다시 실행합니다...")
        ctypes.windll.shell32.ShellExecuteW(
            None, "runas", sys.executable, __file__, None, 1
        )
        sys.exit()

    threading.Thread(target=update_chrome_driver).start()

# 크롬 드라이버 설정
service = Service(CHROME_DRIVER_PATH)
chrome_options = Options()

# 데이터 기억 옵션 추가
chrome_options.add_argument("user-data-dir=" + USER_DATA_DIR)  # 사용자 데이터 디렉토리 설정
chrome_options.add_argument("disable-blink-features=AutomationControlled")

# 비밀번호 저장 변수
password = ""

# 초기 엑셀 파일 저장 함수
def save_to_excel(blog_ids, nicknames, remarks, blog_links, blog_level, visitors, category):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "이름"
    ws['B1'] = "닉네임"
    ws['C1'] = "블로그아이디"
    ws['D1'] = "지수"
    ws['E1'] = "방문자수"
    ws['F1'] = "카테고리"
    ws['G1'] = "비고"
    ws['H1'] = "블로그주소"

    for index, (blog_id, nickname, remark, blog_link) in enumerate(zip(blog_ids, nicknames, remarks, blog_links), start=2):
        ws.cell(row=index, column=1, value="")  # 이름 (비워둠)
        ws.cell(row=index, column=2, value=nickname)
        ws.cell(row=index, column=3, value=blog_id)
        ws.cell(row=index, column=4, value=blog_level[index - 2] if index - 2 < len(blog_level) else "")
        ws.cell(row=index, column=5, value=visitors[index - 2] if index - 2 < len(visitors) else "")
        ws.cell(row=index, column=6, value=category[index - 2] if index - 2 < len(category) else "")
        ws.cell(row=index, column=7, value=remark)
        ws.cell(row=index, column=8, value=blog_link)
        
    # 각 열의 최대 너비를 계산하고 자동으로 너비 설정
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter

        for cell in col:
            try:
                if cell.value:
                    # 셀의 내용을 줄바꿈 기준으로 분리하고, 가장 긴 줄의 길이를 계산
                    cell_lines = str(cell.value).splitlines()
                    max_length = max(max_length, *[len(line) for line in cell_lines])
            except:
                pass

        # Adjust the column width based on max_length (slightly padded for readability)
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # 비고 열(G)의 너비를 강제로 30으로 설정 (필요에 따라 값 조정 가능)
    ws.column_dimensions['G'].width = 20.88

    global filename
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"blog_data_{timestamp}.xlsx"
    
    wb.save(filename)
    logging.info(f"엑셀 파일 '{filename}'에 데이터가 저장되었습니다.")


# 진단 데이터 덮어쓰기 함수
def update_excel_with_diagnosis(blog_level, visitors, category):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    for i, (level, visitor, cat) in enumerate(zip(blog_level, visitors, category), start=2):
        ws.cell(row=i, column=4, value=level)
        ws.cell(row=i, column=5, value=visitor)
        ws.cell(row=i, column=6, value=cat)
    
    # 각 열의 최대 너비를 계산하고 자동으로 너비 설정
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter

        for cell in col:
            try:
                if cell.value:
                    # 셀의 내용을 줄바꿈 기준으로 분리하고, 가장 긴 줄의 길이를 계산
                    cell_lines = str(cell.value).splitlines()
                    max_length = max(max_length, *[len(line) for line in cell_lines])
            except:
                pass

        # Adjust the column width based on max_length (slightly padded for readability)
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

    # 비고 열(G)의 너비를 20.88로 강제 설정
    ws.column_dimensions['G'].width = 20.88
    
    # 비고 열(F)의 너비를 8.75로 강제 설정
    ws.column_dimensions['F'].width = 8.75

    wb.save(filename)
    logging.info(f"엑셀 파일 '{filename}'에 진단 결과가 덮어씌워졌습니다.")
    
# 블로그 ID 추출 함수
def extract_blog_id(url):
    match = re.search(r"blogId=([^&]+)", url)
    if match:
        return match.group(1)
    else:
        return url.rstrip('/').split('/')[-1]
# 크롤링 및 로그인 함수
def crawl_and_login(url):
    driver = webdriver.Chrome(service=service, options=chrome_options)
    try:
        driver.get(url)
        logging.info("URL로 이동 중...")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="password"]')))
        logging.info("비밀번호 입력 필드가 로드되었습니다.")

        text_input = driver.find_element(By.CSS_SELECTOR, 'input[type="password"]')
        text_input.send_keys(password)
        logging.info("비밀번호 입력 완료.")
        text_input.send_keys(Keys.RETURN)
        logging.info("비밀번호 제출 중...")

        time.sleep(5)

        # 나머지 크롤링 코드 유지
        blog_links = []
        blog_ids = []
        nicknames = []
        remarks = []
        blog_level = []
        visitors = []
        category = []

        while True:
            try:
                more_button = driver.find_element(By.CSS_SELECTOR, "button.more_btn#more_ask_li")
                more_button.click()
                time.sleep(2)
            except Exception:
                break

        blog_elements = driver.find_elements(By.CSS_SELECTOR, "a.sns_btn")
        for link_element in blog_elements:
            blog_link = link_element.get_attribute("href")
            blog_links.append(blog_link)
            blog_id = extract_blog_id(blog_link)  # 추출 함수로 블로그 ID 추출
            blog_ids.append(blog_id)

            try:
                nickname_element = link_element.find_element(By.XPATH, "../../..//div[@class='mb_info']")
                full_text = nickname_element.text
                spans = nickname_element.find_elements(By.TAG_NAME, 'span')
                for span in spans:
                    full_text = full_text.replace(span.text, '')
                nickname = full_text.strip()
                nicknames.append(nickname)
            except Exception as e:
                logging.error(f"닉네임 추출 중 오류 발생: {e}")
                nicknames.append("")

            try:
                remark_element = link_element.find_element(By.XPATH, "../../..//div[@class='mb_info']/span[2]")
                remark = remark_element.text
                remarks.append(remark)
            except Exception:
                remarks.append("")

        log_output = "\n".join([f"{i+1}. {blog_id}" for i, blog_id in enumerate(blog_ids)])
        log_text.config(state="normal")
        log_text.delete(1.0, tk.END)
        log_text.insert(tk.END, log_output)
        log_text.config(state="disabled")

        logging.info("블로그 링크가 성공적으로 추출되었습니다.")
        save_to_excel(blog_ids, nicknames, remarks, blog_links, blog_level, visitors, category)
    finally:
        driver.quit()

# 제출 버튼 클릭 시 호출되는 함수


# 연구 시작 함수
def start_research():
    driver = webdriver.Chrome(service=service, options=chrome_options)
    
    try:
        driver.get("https://lablog.co.kr/")
        logging.info("연구 사이트로 이동 중...")
        time.sleep(2)
        
        elements = driver.find_elements(By.CSS_SELECTOR, ".MuiGrid-root.MuiGrid-item.MuiGrid-grid-xs-4.css-1udb513")
        if len(elements) >= 3:
            elements[2].click()
            logging.info("세 번째 버튼 클릭 완료.")
        else:
            logging.error("요소를 찾을 수 없습니다.")
            goto_blog_diagnosis(driver)
            return

        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
        original_window = driver.current_window_handle

        for handle in driver.window_handles:
            if handle != original_window:
                driver.switch_to.window(handle)
                break
        else:
            logging.error("새 창이 열리지 않았습니다.")
            return

        logging.info("새 창으로 전환 완료.")
        time.sleep(2)

        email_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.NAME, "identifier"))
        )
        email_input.send_keys(LOGIN_EMAIL)
        email_input.send_keys(Keys.RETURN)
        logging.info("이메일 입력 및 제출 완료.")

        WebDriverWait(driver, 20).until(EC.number_of_windows_to_be(1))
        driver.switch_to.window(original_window)
        logging.info("새 창이 닫혔습니다. 원래 창으로 돌아왔습니다.")
        time.sleep(2)

        goto_blog_diagnosis(driver)
    
    finally:
        driver.quit()

def goto_blog_diagnosis(driver):
    try:
        # "다시보지않기" 버튼이 있으면 클릭
        dont_show_again_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.XPATH, "//*[text()='다시보지않기']"))
        )
        dont_show_again_button.click()
        logging.info("다시보지않기 버튼 클릭 완료.")
    except Exception:
        logging.info("다시보지않기 버튼이 없습니다. 진행합니다.")

    # "블로그 진단" 버튼을 찾아 클릭
    blog_diagnosis_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//*[text()='블로그 진단']"))
    )
    blog_diagnosis_button.click()
    logging.info("블로그 진단 버튼 클릭 완료.")

    # 입력 필드 초기화 및 값 입력 반복
    input_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, ".MuiInputBase-input.MuiOutlinedInput-input.MuiInputBase-inputSizeSmall.css-2ozrrz"))
    )
    input_field.click()  # 입력 필드 클릭
    logging.info("입력 필드 클릭 완료.")

    # 로그에서 블로그 ID 목록 가져오기
    log_items = log_text.get("1.0", tk.END).strip().split("\n")

    # 데이터를 저장할 리스트
    blog_level = []
    visitors = []
    category = []
    failed_items = []  # 실패한 ID를 저장할 리스트

    # 각 블로그 ID를 입력하고 진단 버튼 클릭 후 데이터 추출
    for item in log_items:
        input_field.click()
        
        # 필드 초기화 (3초 동안 백스페이스 누름)
        start_time = time.time()
        while time.time() - start_time < 3:
            input_field.send_keys(Keys.BACKSPACE)

        # 숫자와 점을 제거하고 순수 ID만 추출
        clean_id = item.split(". ", 1)[1] if ". " in item else item
        
        # 정제된 ID 입력
        input_field.send_keys(clean_id)
        logging.info(f"'{clean_id}' 입력 완료.")

        # 진단 버튼 클릭
        action_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".MuiButtonBase-root.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeMedium.MuiButton-containedSizeMedium.MuiButton-fullWidth.css-9nd1ol"))
        )
        action_button.click()
        logging.info("액션 버튼 클릭 완료.")

        # 로딩바가 사라질 때까지 대기
        WebDriverWait(driver, 60).until_not(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".MuiLinearProgress-root.MuiLinearProgress-colorSecondary.MuiLinearProgress-determinate.css-1hbgb9z"))
        )
        logging.info("진단 완료 후 데이터 추출 준비")

        # 데이터 추출
        try:
            # 지수 추출 부분 수정
            try:
                # 여러 셀렉터 시도
                level_element = None
                selectors = [
                    "text.apexcharts-datalabel",
                    ".apexcharts-datalabel text",
                    "g.apexcharts-datalabel-label text"
                ]
                
                for selector in selectors:
                    try:
                        level_element = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, selector))
                        )
                        if level_element:
                            break
                    except:
                        continue
                        
                if level_element:
                    blog_level_text = level_element.text
                    if ": " in blog_level_text:
                        blog_level_text = blog_level_text.split(": ")[1]
                    elif ":" in blog_level_text:
                        blog_level_text = blog_level_text.split(":")[1]
                    else:
                        blog_level_text = blog_level_text.strip()
                else:
                    blog_level_text = "N/A"
                    
                logging.info(f"추출된 블로그 지수: {blog_level_text}")
                blog_level.append(blog_level_text)
                    
            except Exception as e:
                logging.error(f"지수 추출 중 오류 발생: {e}")
                blog_level.append("N/A")

            # 기존 방문자수와 카테고리 추출 코드는 그대로 유지
            visitor_elements = driver.find_elements(By.CSS_SELECTOR, ".MuiFormControl-root.MuiFormControl-fullWidth.MuiTextField-root.css-ciaeuc")
            if len(visitor_elements) >= 7:
                visitor_value = visitor_elements[6].find_element(By.CSS_SELECTOR, "input").get_attribute("value")
                visitors.append(visitor_value)

            category_elements = driver.find_elements(By.CSS_SELECTOR, ".MuiFormControl-root.MuiFormControl-fullWidth.MuiTextField-root.css-ciaeuc")
            if len(category_elements) >= 2:
                category_value = category_elements[1].find_element(By.CSS_SELECTOR, "input").get_attribute("value")
                category.append(category_value)

        except Exception as e:
            logging.error(f"데이터 추출 중 오류 발생: {e}")
            # 만약 추출이 실패하면 "N/A" 추가하고, 해당 ID를 실패 목록에 추가
            blog_level.append("N/A")
            visitors.append("N/A")
            category.append("N/A")
            failed_items.append(item)

        logging.info(f"지수: {blog_level[-1]}, 방문자수: {visitors[-1]}, 블로그주제: {category[-1]}")

    # 실패한 ID들을 다시 시도
    for item in failed_items:
        logging.info(f"재시도 중: {item}")
        input_field.click()

        # 필드 초기화 (3초 동안 백스페이스 누름)
        start_time = time.time()
        while time.time() - start_time < 3:
            input_field.send_keys(Keys.BACKSPACE)

        # 숫자와 점을 제거하고 순수 ID만 추출
        clean_id = item.split(". ", 1)[1] if ". " in item else item
        
        # 정제된 ID 입력
        input_field.send_keys(clean_id)
        logging.info(f"'{clean_id}' 재시도 입력 완료.")

        # 진단 버튼 클릭
        action_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".MuiButtonBase-root.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeMedium.MuiButton-containedSizeMedium.MuiButton-fullWidth.css-9nd1ol"))
        )
        action_button.click()
        logging.info("재시도 액션 버튼 클릭 완료.")

        # 로딩바가 사라질 때까지 대기
        WebDriverWait(driver, 60).until_not(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".MuiLinearProgress-root.MuiLinearProgress-colorSecondary.MuiLinearProgress-determinate.css-1hbgb9z"))
        )
        logging.info("진단 완료 후 재시도 데이터 추출 준비")

        # 데이터 추출
        try:
            level_element = driver.find_element(By.CSS_SELECTOR, "text.apexcharts-datalabel")
            blog_level_text = level_element.text.split(": ")[1] if ": " in level_element.text else "N/A"
            blog_level[log_items.index(item)] = blog_level_text

            visitor_elements = driver.find_elements(By.CSS_SELECTOR, ".MuiFormControl-root.MuiFormControl-fullWidth.MuiTextField-root.css-ciaeuc")
            if len(visitor_elements) >= 7:
                visitor_value = visitor_elements[6].find_element(By.CSS_SELECTOR, "input").get_attribute("value")
                visitors[log_items.index(item)] = visitor_value

            category_elements = driver.find_elements(By.CSS_SELECTOR, ".MuiFormControl-root.MuiFormControl-fullWidth.MuiTextField-root.css-ciaeuc")
            if len(category_elements) >= 2:
                category_value = category_elements[1].find_element(By.CSS_SELECTOR, "input").get_attribute("value")
                category[log_items.index(item)] = category_value

        except Exception as e:
            logging.error(f"재시도 데이터 추출 중 오류 발생: {e}")
            blog_level[log_items.index(item)] = "N/A"
            visitors[log_items.index(item)] = "N/A"
            category[log_items.index(item)] = "N/A"

    # 추출된 데이터 업데이트
    update_excel_with_diagnosis(blog_level, visitors, category)
def on_submit():
    global password
    password = password_entry.get()
    url = url_entry.get()
    threading.Thread(target=crawl_and_login, args=(url,)).start()

# 연구 시작 스레드 함수
def on_start_research():
    threading.Thread(target=start_research).start()
    
# GUI 설정
root = tk.Tk()
root.title("크롤링 프로그램")
root.geometry("550x650")
root.configure(bg="#f0f0f0")

# 창 크기 고정
root.resizable(width=False, height=False)

title_font = font.Font(family="Helvetica", size=16, weight="bold")
label_font = font.Font(family="Helvetica", size=12)
button_font = font.Font(family="Helvetica", size=12, weight="bold")

title_label = tk.Label(root, text="크롤링 프로그램", font=title_font, bg="#f0f0f0", fg="#333")
title_label.pack(pady=20)

url_label = tk.Label(root, text="접속할 링크를 입력하세요:", font=label_font, bg="#f0f0f0", fg="#333")
url_label.pack(pady=5)

url_entry = tk.Entry(root, width=40, font=label_font)
url_entry.pack(pady=5)

password_label = tk.Label(root, text="비밀번호를 입력하세요:", font=label_font, bg="#f0f0f0", fg="#333")
password_label.pack(pady=5)

password_entry = tk.Entry(root, width=40, font=label_font, show="*")
password_entry.pack(pady=5)

submit_button = tk.Button(root, text="제출", command=on_submit, font=button_font, bg="#4CAF50", fg="white", padx=10, pady=5)
submit_button.pack(pady=20)

update_driver_button = tk.Button(root, text="크롬 드라이버 업데이트", command=on_update_driver, font=button_font, bg="#FF5733", fg="white", padx=10, pady=5)
update_driver_button.pack(pady=10)

log_label = tk.Label(root, text="추출한 블로그ID:", font=label_font, bg="#f0f0f0", fg="#333")
log_label.pack(pady=5)

# 로그 출력 영역 (수정 불가)
log_text = scrolledtext.ScrolledText(root, width=58, height=10, font=label_font, state="disabled")
log_text.pack(pady=5)

research_button = tk.Button(root, text="연구시작", command=on_start_research, font=button_font, bg="#FF5733", fg="white", padx=10, pady=5)
research_button.pack(pady=20)

root.mainloop()