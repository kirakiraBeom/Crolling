# 해당 파일을 from coupanq.ExcelConverter import EXCELConverter 등의 형태로 import 해서 EXCELConverter.run 형태로 외부에서 실행

from typing import Callable
import pandas as pd
import json
from tkinter import Tk, filedialog, messagebox
import os
from openpyxl import load_workbook
import re
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

JSON_KEYWORDS_PATH = os.getenv(
    "JSON_KEYWORDS_PATH", "C:/Temp/JSON/JSON/additionalKeywords.json"
)
JSON_MAPPING_PATH = os.getenv("JSON_MAPPING_PATH", "C:/Temp/JSON/JSON/keywords.json")


def setup_tk_window_for_taskbar() -> Tk:
    """
    Tkinter 창을 작업 표시줄에만 표시되도록 설정합니다.

    Returns:
        Tk: 설정된 Tkinter root 창 객체.
    """
    root = Tk()
    root.withdraw()  # 기본 창 숨김
    root.attributes("-alpha", 0)  # 창 투명도 설정 (완전히 투명하게)
    root.iconify()  # 최소화하여 작업 표시줄에 표시
    return root


def select_excel_file() -> str:
    """
    파일 열기 대화상자를 통해 엑셀 파일을 선택합니다.

    Returns:
        str: 선택된 파일의 경로. 선택되지 않으면 빈 문자열을 반환합니다.
    """
    file_path = filedialog.askopenfilename(
        title="엑셀 파일 선택",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
    )

    if not file_path:
        messagebox.showwarning("파일 선택", "엑셀 파일이 선택되지 않았습니다.")

    return file_path


def load_mapping_json() -> dict:
    """
    JSON 파일에서 매핑 데이터를 로드합니다.

    Returns:
        dict: 매핑 데이터를 포함하는 딕셔너리.
              파일이 없거나 오류가 발생한 경우 None을 반환합니다.
    """
    if not os.path.exists(JSON_MAPPING_PATH):
        messagebox.showwarning(
            "파일 없음", f"JSON 파일 '{JSON_MAPPING_PATH}'이(가) 존재하지 않습니다."
        )
        return None

    try:
        with open(JSON_MAPPING_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        # comments 키 제거
        return {k: v for k, v in data.items() if k != "comments"}
    except Exception as e:
        messagebox.showerror(
            "오류", f"JSON 파일을 로드하는 중 오류가 발생했습니다: {str(e)}"
        )
        return None


def load_keywords_from_json() -> (
    tuple[list[any], dict[str, list[any]], dict[str, str], dict[str, list[str]]]
):
    """
    JSON 파일에서 추가 키워드, 건너뛸 키워드 매핑, 추가 매핑, 강제 키워드 매핑을 로드합니다.

    Returns:
        Tuple[
            List[Any],                       # additional_keywords
            Dict[str, List[Any]],           # skip_keywords_mapping
            Dict[str, str],                 # additional_mappings
            Dict[str, List[str]]            # enforce_keywords_mapping
        ]
    Raises:
        ValueError: JSON 파일을 로드하거나 처리하는 중 오류 발생 시.
    """
    try:
        with open(JSON_KEYWORDS_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)

        # 정규식 키워드 처리
        additional_keywords = [
            (
                re.compile(keyword["regex"])
                if isinstance(keyword, dict) and "regex" in keyword
                else keyword
            )
            for keyword in data.get("additional_keywords", [])
        ]

        # 건너뛸 키워드 매핑
        skip_keywords_mapping = data.get("skip_keywords_mapping", {})

        additional_mappings = data.get("additional_mappings", {})

        enforce_keywords_mapping = data.get("enforce_keywords_mapping", {})

        return (
            additional_keywords,
            skip_keywords_mapping,
            additional_mappings,
            enforce_keywords_mapping,
        )
    except FileNotFoundError:
        raise ValueError(f"JSON 파일 '{JSON_KEYWORDS_PATH}'이(가) 존재하지 않습니다.")
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON 파일 파싱 실패: {str(e)}")
    except Exception as e:
        raise ValueError(f"JSON 파일 로드 중 알 수 없는 오류 발생: {str(e)}")


def validate_file_and_mapping(root: Tk) -> tuple[str, dict]:
    """
    파일 선택 및 JSON 매핑 데이터를 로드합니다.

    Args:
        root (Tk): Tkinter root 창 객체.

    Returns:
        tuple[str, dict]: 선택된 파일 경로와 매핑 데이터 딕셔너리의 튜플.
                          선택이 취소되거나 매핑 데이터가 없는 경우 (None, None)를 반환.
    """
    # 파일 선택
    file_path = select_excel_file()
    if not file_path:
        messagebox.showwarning("파일 선택", "파일이 선택되지 않았습니다.")
        root.destroy()
        return None, None

    # 매핑 로드
    mapping = load_mapping_json()
    if not mapping:
        root.destroy()
        return None, None

    return file_path, mapping


def check_header(df: pd.DataFrame, header_name: str, root: Tk) -> bool:
    """
    데이터프레임에 특정 헤더가 존재하는지 확인합니다.

    Args:
        df (pd.DataFrame): 데이터프레임.
        header_name (str): 확인할 헤더 이름.
        root (Tk): Tkinter root 창 객체.

    Returns:
        bool: 헤더가 존재하면 True, 그렇지 않으면 False.
    """
    if header_name not in df.columns:
        messagebox.showerror(
            "헤더 없음", f"헤더 '{header_name}'이(가) 파일에 존재하지 않습니다."
        )
        root.destroy()
        raise ValueError(
            f"헤더 '{header_name}'이(가) 데이터프레임에 존재하지 않습니다."
        )
    return True


def process_keywords_with_skip_and_enforce(
    original_value: str,
    mapped_value: str,
    additional_keywords: list,
    skip_keywords_mapping: dict,
    enforce_keywords_mapping: dict,
    additional_texts: list,
) -> list:
    """
    추가 키워드와 강제 키워드를 처리하며, 스킵 조건을 적용.

    Args:
        original_value (str): 원본 데이터 값.
        mapped_value (str): 매핑된 데이터 값.
        additional_keywords (list): 추가 키워드 목록.
        skip_keywords_mapping (dict): 스킵 키워드 매핑.
        enforce_keywords_mapping (dict): 강제 키워드 매핑.
        additional_texts (list): 추가 키워드를 저장할 리스트.

    Returns:
        list: 최종 추가 텍스트 리스트.
    """
    # 스킵 키워드와 강제 키워드 가져오기
    skip_keywords = skip_keywords_mapping.get(mapped_value, [])
    enforce_keywords = enforce_keywords_mapping.get(mapped_value, [])

    # 추가 키워드 처리
    for keyword in additional_keywords:
        if isinstance(keyword, re.Pattern):  # 정규식 처리
            match = keyword.search(str(original_value))
            if match and not should_skip(match.group(), skip_keywords):
                additional_texts.append(match.group())
        elif keyword in str(original_value):  # 일반 키워드 처리
            if keyword not in skip_keywords:
                additional_texts.append(keyword)

    # 강제 키워드 처리
    for key, enforced_keywords in enforce_keywords_mapping.items():
        if key in str(
            mapped_value
        ):  # 매핑된 데이터에 강제 키워드의 키가 포함되는지 확인
            for enforced_keyword in enforced_keywords:
                if (
                    enforced_keyword in str(original_value)
                    and enforced_keyword not in additional_texts
                    and not should_skip(enforced_keyword, skip_keywords)
                ):
                    additional_texts.append(enforced_keyword)

    return additional_texts


def should_skip(value: str, skip_keywords: list) -> bool:
    """
    주어진 값이 스킵 키워드에 포함되는지 확인.

    Args:
        value (str): 확인할 값.
        skip_keywords (list): 스킵 키워드 목록 (정규식 또는 일반 문자열).

    Returns:
        bool: 스킵 키워드에 포함되면 True, 아니면 False.
    """
    for skip_keyword in skip_keywords:
        if isinstance(skip_keyword, dict) and "regex" in skip_keyword:
            skip_regex = re.compile(skip_keyword["regex"])
            if skip_regex.fullmatch(value):
                return True
        elif value == skip_keyword:
            return True
    return False


def process_mapped_data(
    df,
    header_name: str,
    mapping: dict,
    additional_keywords: list,
    skip_keywords_mapping: dict,
    enforce_keywords_mapping: dict,
    additional_mappings: dict,
) -> tuple[list, list]:
    """
    매핑된 데이터를 처리합니다.

    Args:
        df (pd.DataFrame): 입력 데이터프레임.
        header_name (str): 처리할 헤더 이름.
        mapping (dict): 매핑 데이터.
        additional_keywords (list): 추가 키워드 목록.
        skip_keywords_mapping (dict): 스킵 키워드 매핑 데이터.
        enforce_keywords_mapping (dict): 강제 키워드 매핑 데이터.
        additional_mappings (dict): 추가 매핑 데이터.

    Returns:
        tuple[list, list]: 원본 값 리스트와 매핑된 값 리스트의 튜플.
    """
    original_column = []  # 원본 옵션명
    mapped_column = []  # 수정된 옵션명

    for _, row in df.iterrows():
        original_value = row[header_name]
        mapped_value = original_value  # 초기값은 원래 값

        # 매핑된 값 확인 및 설정
        for key, value in mapping.items():
            if key in str(original_value):
                mapped_value = value
                break

        # 추가 키워드 확인 및 첨부
        additional_texts = []
        additional_texts = process_keywords_with_skip_and_enforce(
            original_value,
            mapped_value,
            additional_keywords,
            skip_keywords_mapping,
            enforce_keywords_mapping,
            additional_texts,
        )

        # JSON에서 추가 매핑 처리
        for keyword, extra_keyword in additional_mappings.items():
            if keyword in str(original_value) and extra_keyword not in additional_texts:
                additional_texts.append(extra_keyword)

        # 매핑된 값을 추가 키워드와 결합
        if additional_texts:
            mapped_value += f" {', '.join(additional_texts)}"

        original_column.append(original_value)
        mapped_column.append(mapped_value)

    return original_column, mapped_column


def create_column_if_missing_and_add_to_result(
    df: pd.DataFrame,
    final_result_df: pd.DataFrame,
    new_column_name: str,
    required_columns: list[str],
    calculation_func: Callable[[pd.DataFrame], pd.Series],
    success_message: str,
    error_message: str,
) -> bool:
    """
    데이터프레임에 특정 열이 없으면 생성하고, 결과 데이터프레임에 추가.

    Args:
        df (pd.DataFrame): 원본 데이터프레임.
        final_result_df (pd.DataFrame): 결과 데이터프레임.
        new_column_name (str): 생성하거나 추가할 열 이름.
        required_columns (list): 계산에 필요한 열 이름들.
        calculation_func (Callable): 계산 로직을 담은 함수.
        success_message (str): 열 생성 성공 시 출력할 메시지.
        error_message (str): 열 생성 실패 시 출력할 메시지.

    Returns:
        bool: 열이 성공적으로 생성되거나 이미 존재하는 경우 True, 실패하면 False.
    """
    if new_column_name not in df.columns:
        if all(col in df.columns for col in required_columns):
            # 열 생성
            try:
                df[new_column_name] = calculation_func(df)
                print(success_message)
            except Exception as e:
                print(f"열 생성 중 오류 발생: {str(e)}")
                return False
        else:
            print(error_message)
            return False

    # 결과 데이터프레임에 열 추가
    if new_column_name not in final_result_df.columns:
        final_result_df[new_column_name] = df[new_column_name]

    return True


def save_and_format_excel(
    df: pd.DataFrame, save_path: str, hidden_columns: list = None
) -> None:
    """
    데이터프레임을 엑셀로 저장하고 서식을 설정.

    Args:
        df (pd.DataFrame): 저장할 데이터프레임.
        save_path (str): 저장 경로.
        hidden_columns (list): 숨길 열 이름 리스트.

    Returns:
        None
    """
    df.to_excel(save_path, index=False, engine="openpyxl")
    wb = load_workbook(save_path)
    ws = wb.active

    # 열 너비 자동 조정
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        col_name = ws.cell(row=1, column=col_idx).value
        if col_name:
            max_length = max(
                (len(str(cell.value)) for cell in column_cells if cell.value), default=0
            )
            adjusted_width = max_length * 1.3 + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 숨김 처리
    if hidden_columns:
        for col_name in hidden_columns:
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                ws.column_dimensions[get_column_letter(col_idx)].hidden = True
            else:
                print(f"열 '{col_name}'이 데이터프레임에 없습니다. 숨김을 건너뜁니다.")

    # 필터 적용
    if not df.empty and df.shape[1] > 0:
        last_column = ws.cell(row=1, column=df.shape[1]).column_letter
        ws.auto_filter.ref = f"A1:{last_column}1"

    # A열 오른쪽 정렬
    for cell in ws["A"]:  # "A"열 전체 순회
        cell.alignment = Alignment(horizontal="right")

    # 틀 고정
    ws.freeze_panes = "A2"

    # 파일 저장
    try:
        wb.save(save_path)
        print(f"파일이 저장되었습니다: {save_path}")
    except PermissionError:
        print("파일 저장에 실패했습니다. 파일이 열려 있는지 확인하세요.")
    except Exception as e:
        print(f"파일 저장 중 오류가 발생했습니다: {str(e)}")


def extract_and_save_options_with_mapping():
    """
    특정 헤더의 값을 추출하고 JSON 매핑을 적용하여 엑셀 파일로 저장합니다.
    """

    # Tkinter 창 설정
    root = setup_tk_window_for_taskbar()

    # 파일 및 매핑 데이터 확인
    file_path, mapping = validate_file_and_mapping(root)
    if not file_path or not mapping:
        return

    try:
        # 엑셀 파일 읽기
        try:
            df = pd.read_excel(file_path, engine="openpyxl")
        except FileNotFoundError:
            messagebox.showerror("오류", "파일을 찾을 수 없습니다.")
            root.destroy()
            return
        except Exception as e:
            messagebox.showerror(
                "오류", f"엑셀 파일을 읽는 중 오류가 발생했습니다: {str(e)}"
            )
            root.destroy()
            return

        header_name = "옵션명"  # 추출할 헤더 이름

        # 헤더 확인
        if not check_header(df, header_name, root):
            return

        # 키워드 로드
        (
            additional_keywords,
            skip_keywords_mapping,
            additional_mappings,
            enforce_keywords_mapping,
        ) = load_keywords_from_json()

        # 맵핑된 데이터 처리
        original_column, mapped_column = process_mapped_data(
            df,
            header_name="옵션명",
            mapping=mapping,
            additional_keywords=additional_keywords,
            skip_keywords_mapping=skip_keywords_mapping,
            enforce_keywords_mapping=enforce_keywords_mapping,
            additional_mappings=additional_mappings,
        )

        # 새로운 데이터프레임 생성
        result_df = pd.DataFrame(
            {f"{header_name}_원본": original_column, f"{header_name}": mapped_column}
        )

        # 결과 데이터프레임 생성
        final_result_df = result_df.copy()

        # 순 판매 금액 처리
        if create_column_if_missing_and_add_to_result(
            df,
            final_result_df,
            "순 판매 금액(전체 거래 금액 - 취소 금액)",
            ["전체 거래 금액", "취소 금액"],
            lambda df: df["전체 거래 금액"] - df["취소 금액"],
            "순 판매 금액(전체 거래 금액 - 취소 금액) 열이 생성되었습니다.",
            "전체 거래 금액 또는 취소 금액 열이 없어 순 판매 금액을 계산할 수 없습니다.",
        ):
            # 간략화된 열 이름 추가
            final_result_df.rename(
                columns={"순 판매 금액(전체 거래 금액 - 취소 금액)": "순 판매 금액"},
                inplace=True,
            )

        # 순 판매 상품 수 처리
        if create_column_if_missing_and_add_to_result(
            df,
            final_result_df,
            "순 판매 상품 수(전체 거래 상품 수 - 취소 상품 수)",
            ["전체 거래 상품 수", "취소 상품 수"],
            lambda df: df["전체 거래 상품 수"] - df["취소 상품 수"],
            "순 판매 상품 수(전체 거래 상품 수 - 취소 상품 수) 열이 생성되었습니다.",
            "전체 거래 상품 수 또는 취소 상품 수 열이 없어 순 판매 상품 수를 계산할 수 없습니다.",
        ):
            # 간략화된 열 이름 추가
            final_result_df.rename(
                columns={
                    "순 판매 상품 수(전체 거래 상품 수 - 취소 상품 수)": "순 판매 상품 수"
                },
                inplace=True,
            )

        # 사용자에게 출력 여부 확인
        include_zero = messagebox.askyesno(
            "출력 여부", "순 판매 상품 수가 0인 데이터를 출력하시겠습니까?"
        )

        sort_or_not = messagebox.askyesno(
            "정렬 여부", "옵션명을 기준으로 오름차순 정렬을 진행하시겠습니까?"
        )

        if not include_zero:
            df = df[df["순 판매 상품 수(전체 거래 상품 수 - 취소 상품 수)"] > 0]

        # 순 판매 상품 수가 0인 행 제거
        if not include_zero and "순 판매 상품 수" in final_result_df.columns:
            final_result_df = final_result_df[final_result_df["순 판매 상품 수"] > 0]

        # 정렬: 옵션명을 기준으로 오름차순 정렬
        if sort_or_not:  # 사용자가 "예"를 선택하면 오름차순 정렬
            final_result_df.sort_values(
                by=["옵션명"],  # 정렬 기준 열
                ascending=[True],  # 오름차순 정렬
                inplace=True,
            )

        # 파일 저장
        base_name = os.path.splitext(os.path.basename(file_path))[0]
        default_filename = f"mapped_extracted_{base_name}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="저장할 파일 경로 선택",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            initialfile=default_filename,
        )

        # 파일 저장 및 서식 적용
        if not save_path:
            messagebox.showwarning("저장 취소", "파일 저장이 취소되었습니다.")
        else:
            save_and_format_excel(
                final_result_df, save_path, hidden_columns=[f"{header_name}_원본"]
            )
            messagebox.showinfo("성공", f"매핑된 데이터가 저장되었습니다: {save_path}")

    except Exception as e:
        messagebox.showerror("오류", f"데이터 처리 중 오류가 발생했습니다: {str(e)}")

    finally:
        root.destroy()  # 창 닫기
        return


# 프로그램 실행
class EXCELConverter:
    @staticmethod
    def run():
        extract_and_save_options_with_mapping()
