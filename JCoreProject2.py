import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import datetime
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# "차종/연식/색상" 필드에서 대체할 키워드와 대체할 값의 딕셔너리
replace_list = {
    "부재 시 문 앞에 놓아주세요.": "",
    "배송 전에 미리 연락바랍니다.": "",
    "배송 전 미리 연락바랍니다.": "",
    "배송 전 미리 연락해 주세요":"",
    "부재 시 경비실에 맡겨주세요.": "",
    "빠른 배송 부탁드립니다.": "",
    "택배함에 보관해 주세요.": "",
    "[추가옵션]": "",
    "테두리 엣지형 변경=": "",
    "테두리 엣지형 변경:":"",
    "트렁크매트 추가:": "",
    "차종:": "/",
    "구성:" : "/",
    "색상:" : "/",
    "3열추가:" : "",
    "필수!!" : "",
    "차량의 유종":"",
    r"(안내 참고):" : "",
    ",": "",
    "차종명=": "/",
    r"매트구성(승용)=": "/",
    "매트색상=": "/",
    "매트색상:": "/",
    "해당 옵션 클릭 후 하단에 제품명을 입력": "",
    r"매트구성 (승용):":"/",
    "^":"",
    r"공지 차종표 확인▶" : "",
    r"공지 차종표 확인 ▶":"",
    "세부차종":"",
    "옵션":"",
    r"★연식★:" : "/",
    r"트렁크 구매시 세부 차종 (해당없으면 X):":"/",
    r"트렁크 구매시  (해당없으면 X)을 입력해주세요.=":"/",
    r"을 입력해주세요.=":"",
    ";":"",
    ":":"",
}

def select_first_file():
    global first_file_path
    first_file_path = filedialog.askopenfilename(title="첫 번째 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
    if first_file_path:
        messagebox.showinfo("파일 선택", f"첫 번째 파일이 선택되었습니다: {first_file_path}")
        print(f"첫 번째 파일 경로: {first_file_path}")

def select_second_file():
    global second_file_path
    second_file_path = filedialog.askopenfilename(title="두 번째 파일 선택", filetypes=[("Excel files", "*.xlsx *.xls")])
    
    if first_file_path and second_file_path:
        process_files(first_file_path, second_file_path)

def add_delivery_message(row):
    """'쇼핑몰'이 '센트럴리빙' 또는 '현대이지웰'인 경우 '배송메시지'를 앞에 추가"""
    # 배송메시지가 비어있는 경우 빈 문자열로 설정
    delivery_message = row['배송메시지'] if pd.notna(row['배송메시지']) else ""
    
    if row['쇼핑몰'] in ['센트럴리빙', '현대이지웰']:
        return f"{delivery_message} {row['옵션']}".strip()  # 앞뒤 공백 제거
    return row['옵션']

def append_order_quantity(row, base_text):
    """주문수량이 2개 이상인 경우 텍스트 끝에 주문수량을 추가"""
    if row['주문수량'] >= 2:
        return f"{base_text} ★{int(row['주문수량'])}개★"
    return base_text

def replace_items_in_field(text):
    """차종/연식/색상 텍스트에서 replace_list에 있는 항목을 대체하고 연속된 '/'와 공백을 하나로 줄임"""
    if not isinstance(text, str):  # text가 문자열인지 확인
        return text  # 문자열이 아니면 그대로 반환
    for old, new in replace_list.items():
        text = text.replace(old, new)
    # 숫자 뒤에 마침표가 오는 형태 (예: 04.)를 제거
    text = re.sub(r'\b\d+\.\s*', '', text)
    # (+숫자원) 형태를 제거
    text = re.sub(r'\(\+\d+원\)', '', text)
    # 연속된 '/'와 공백을 하나로 줄임, 그리고 '/' 전후로 공백을 하나씩 추가
    text = re.sub(r'(\s*/\s*)+', ' / ', text)
    # 맨 앞의 '/' 제거
    
    # "(안내 참고):" 패턴 삭제 (정규 표현식 사용)
    text = re.sub(r"\(안내 참고\):\s*", "", text)  # "(안내 참고):"와 그 뒤의 공백 삭제
    
    text = text.lstrip(' /')
    return text.strip()

def process_files(first_file, second_file):
    # Excel 파일 불러오기
    df1 = pd.read_excel(first_file)
    df2 = pd.read_excel(second_file)

    # 첫 번째 파일의 특정 열을 두 번째 파일에 매핑
    df2['성명'] = df1['받는분성함']
    df2['차종/연식/색상'] = df1.apply(
        lambda row: append_order_quantity(row, add_delivery_message(row)),
        axis=1
    )
    # "차종/연식/색상" 필드에서 항목 대체 및 '/' 전후로 공백 처리
    df2['차종/연식/색상'] = df2['차종/연식/색상'].apply(replace_items_in_field)
    
    df2['주소'] = df1['받는분주소']
    df2['연락처'] = df1['반는분연락처']
    df2['기타연락처'] = df1['받는분기타연락처']
    df2['배송메세지'] = df1['배송메시지']

    # 두 번째 파일의 디렉토리 가져오기
    second_file_dir = os.path.dirname(second_file)

    # 저장할 파일 이름과 경로 설정
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    base_output_filename = f"코일매트_{current_date}.xlsx"
    output_filename = os.path.join(second_file_dir, base_output_filename)

    # 파일 이름 중복 시 숫자 추가
    counter = 1
    while os.path.exists(output_filename):
        output_filename = os.path.join(second_file_dir, f"코일매트_{current_date}_{counter}.xlsx")
        counter += 1

    # 수정된 파일 저장
    df2.to_excel(output_filename, index=False)

    # 워크북 불러오기 및 활성 시트 접근
    workbook = load_workbook(output_filename)
    sheet = workbook.active

    # 셀 색상과 테두리 스타일 정의
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    blue_fill = PatternFill(start_color="B7CFE8", end_color="B7CFE8", fill_type="solid")
    highlight_fill = PatternFill(start_color="66CCFF", end_color="66CCFF", fill_type="solid")  # 추출된 색상
    border_style = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    alignment_style = Alignment(wrap_text=True)  # 자동 줄바꿈 설정

    # 폰트 스타일 정의
    header_font = Font(name="맑은 고딕", size=11, bold=True)  # 1행 폰트 스타일
    body_font = Font(name="Arial", size=10)  # 나머지 셀 폰트 스타일

    # 각 열의 너비 설정
    column_widths = {
        'A': 10,  # 날짜
        'B': 15,  # 성명
        'C': 17,  # 유형
        'D': 50,  # 차종/연식/색상
        'E': 10,  # 수량
        'F': 53,  # 주소
        'G': 15,  # 연락처
        'H': 15,  # 기타연락처
        'I': 38,  # 배송메세지
        'J': 13   # 한진택배  
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # 첫 번째 행의 높이 설정
    sheet.row_dimensions[1].height = 25

    # 첫 번째 행의 지정된 셀에 색상, 테두리 및 폰트 적용
    for row in range(1, 2):  # 1행에 적용
        for col in ['A', 'B', 'E', 'F', 'G', 'H', 'I', 'J']:  # 회색 적용
            cell = sheet[f"{col}{row}"]
            cell.fill = gray_fill
            cell.alignment = alignment_style  # 자동 줄바꿈
            cell.font = header_font  # 1행 폰트 스타일
            cell.border = border_style  # 모든 테두리 추가
        for col in ['C', 'D']:  # 파란색 적용
            cell = sheet[f"{col}{row}"]
            cell.fill = blue_fill
            cell.alignment = alignment_style  # 자동 줄바꿈
            cell.font = header_font  # 1행 폰트 스타일
            cell.border = border_style  # 모든 테두리 추가

    # 나머지 셀에 폰트 적용 및 테두리 추가
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):
        for cell in row: 
            cell.font = body_font  # 나머지 셀 폰트 스타일
            cell.alignment = alignment_style  # 자동 줄바꿈
            if cell.value is not None:  # 데이터가 존재할 경우
                cell.border = border_style  # 모든 테두리 추가

    # "주소" 필드가 중복된 행을 찾아 강조 표시 및 테두리 적용
    duplicate_addresses = df2['주소'][df2['주소'].duplicated(keep=False)]
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=9):  # I열까지 적용
        if row[5].value in duplicate_addresses.values:  # '주소'는 F열 (index 5) 기준
            for cell in row:
                cell.fill = highlight_fill
                cell.alignment = alignment_style  # 자동 줄바꿈
                cell.border = border_style  # 모든 테두리 추가

    # 업데이트된 서식을 포함하여 워크북 저장
    workbook.save(output_filename)
    messagebox.showinfo("완료", f"파일이 저장되었습니다: {output_filename}")
    print(f"파일이 저장되었습니다: {output_filename}")

root = tk.Tk()
root.title("J CORE PROJECT")
root.geometry("400x250")
root.configure(bg="#2c3e50")

label = tk.Label(root, text="엑셀 파일을 선택하세요:", font=("Arial", 16), bg="#2c3e50", fg="white")
label.pack(pady=20)

first_button = tk.Button(root, text="첫 번째 파일 선택", font=("Arial", 14), bg="#3498db", fg="white", command=select_first_file)
first_button.pack(pady=10)

second_button = tk.Button(root, text="두 번째 파일 선택", font=("Arial", 14), bg="#e74c3c", fg="white", command=select_second_file)
second_button.pack(pady=10)

root.mainloop()
